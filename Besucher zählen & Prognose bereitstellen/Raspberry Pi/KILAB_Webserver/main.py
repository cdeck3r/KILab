import pygal, lxml
from flask import Flask, Response
from openpyxl import load_workbook
from datetime import date, datetime
workbook = load_workbook(filename="maerz-2021.xlsx")

app = Flask(__name__)


@app.route('/')
def index():
    """ render svg on html """
    return """
<html>
    <body>    
         <figure>
             <embed style='width:75%;' class='img-responsive col-xs-12' type="image/svg+xml" src="/graph/" />
            </figure>
    </body>
</html>
"""


@app.route('/graph/')
def graph():
    timeblock = []
    sheet = workbook.active
    today = date.today()

    # dd/mm/YY
    d1 = today.strftime("%d/%m/%Y")
    print(d1)
    dayofmonth = today.strftime("%d")
    dayofmonth = int(dayofmonth.rstrip('0'))

    now = datetime.now()
    current_time = now.strftime("%H:%M")
    print("Current Time =", current_time)

    for value in sheet.iter_rows(min_row=1, max_row=31,
                                 min_col=1, max_col=3,
                                 values_only=True):
        timeblock.append(value)

    for list in timeblock:
        # print(list)
        pass

    # Set data for today
    print("Today: " + sheet.cell(row=dayofmonth + 1, column=3).value)
    todaylist = sheet.cell(row=dayofmonth + 1, column=3).value

    #  Converts the entry to int list
    l = todaylist.split(',')
    l = [int(i) for i in l]

    line_chart = pygal.Bar()
    line_chart.title = 'Stoßzeiten KILAB'
    #  line_chart.x_labels = 'Montag', 'Dienstag', 'Mittwoch', 'Donnerstag', 'Freitag', 'Samstag'
    line_chart.x_labels = '08:00 - 09:30', '09:45 - 11:15', '11:30 - 13:00', '13:45 - 15:15', '15:30 - 17:00', \
                          '17:15 - 18:45', '19:00 - 20:30'

    line_chart.add('Aktuell', [0, 0, 0, 3, 0])
    line_chart.add('Prognose', l)
    line_chart.render_to_file('dice_visual.svg')
    return Response(response=line_chart.render(), content_type='image/svg+xml')


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5050)